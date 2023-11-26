from openpyxl import Workbook,load_workbook
wb = load_workbook(filename="abc.xlsx")
sh = wb.active

row_R = sh.max_row # for maximum rows
row_C = sh.max_column #for maximun columns

for i in range(1 ,row_R+1):
    for j in range(1,row_C+1):
        print(sh.cell(row=i, column =j).value,end=" ")
    print("\n")

#print(sh['A3'].value)
#print(row_C ,row_R)